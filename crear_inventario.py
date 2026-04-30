import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

wb = Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", start_color=hex_color.lstrip("#"), end_color=hex_color.lstrip("#"))

def apply_border(cell, style="thin"):
    s = Side(style=style)
    cell.border = Border(left=s, right=s, top=s, bottom=s)

# ── DATA ─────────────────────────────────────────────────────────────────────
categories = [
    {
        "name": "MANTENIMIENTO PREVENTIVO",
        "bg": "DBEAFE", "fg": "1E40AF",
        "items": [
            ("FIL-ACE-TOY","Filtro de aceite Toyota","Toyota",350,5,3,"Unidad"),
            ("FIL-ACE-HON","Filtro de aceite Honda","Honda",350,5,3,"Unidad"),
            ("FIL-ACE-MAZ","Filtro de aceite Mazda","Mazda",350,5,3,"Unidad"),
            ("FIL-ACE-NIS","Filtro de aceite Nissan","Nissan",350,5,3,"Unidad"),
            ("FIL-ACE-CHE","Filtro de aceite Chevrolet","Chevrolet",350,5,3,"Unidad"),
            ("FIL-ACE-HYU","Filtro de aceite Hyundai","Hyundai",350,5,3,"Unidad"),
            ("FIL-ACE-KIA","Filtro de aceite Kia","Kia",350,5,3,"Unidad"),
            ("FIL-ACE-FOR","Filtro de aceite Ford","Ford",350,5,3,"Unidad"),
            ("FIL-AIR-GEN","Filtro de aire de motor","Todas",450,6,3,"Unidad"),
            ("FIL-CAB-GEN","Filtro de cabina / habitáculo","Todas",650,4,2,"Unidad"),
            ("FIL-COM-GEN","Filtro de combustible","Todas",750,4,2,"Unidad"),
            ("FIL-ATF-GEN","Filtro de transmisión automática","Todas",900,3,2,"Unidad"),
            ("ACE-5W30-SIN","Aceite 5W-30 sintético (cuarto)","Todas",550,20,10,"Cuarto"),
            ("ACE-10W30-SEM","Aceite 10W-30 semi-sintético (cuarto)","Todas",450,20,10,"Cuarto"),
            ("ACE-0W20-SIN","Aceite 0W-20 sintético (cuarto)","Toyota/Honda",650,10,5,"Cuarto"),
            ("ACE-ATF-DEX","Aceite ATF Dexron III/VI","Todas",500,8,4,"Cuarto"),
            ("ACE-MAN-75W90","Aceite transmisión manual 75W-90","Todas",550,5,2,"Cuarto"),
            ("LIQ-FRE-DOT3","Líquido de frenos DOT 3","Todas",350,8,4,"Unidad"),
            ("LIQ-FRE-DOT4","Líquido de frenos DOT 4","Todas",400,8,4,"Unidad"),
            ("LIQ-REF-VER","Líquido refrigerante verde","Todas",500,6,3,"Unidad"),
            ("LIQ-REF-ROS","Líquido refrigerante rosado","Toyota/Honda",600,6,3,"Unidad"),
            ("LIQ-DIR-HID","Líquido dirección hidráulica","Todas",400,5,3,"Unidad"),
            ("LIQ-LIM-PAR","Líquido limpiaparabrisas","Todas",200,10,5,"Unidad"),
            ("BUJ-NGK-STD","Bujías NGK estándar (set x4)","Todas",1200,6,3,"Set"),
            ("BUJ-NGK-IRI","Bujías NGK iridio (set x4)","Todas",2800,4,2,"Set"),
            ("BUJ-DEN-STD","Bujías Denso (set x4)","Todas",1300,4,2,"Set"),
            ("CAB-BUJ-GEN","Cables de bujías (juego)","Todas",1800,3,2,"Juego"),
            ("BOB-ENC-LAP","Bobina de encendido tipo lápiz","Todas",2500,4,2,"Unidad"),
        ]
    },
    {
        "name": "SISTEMA DE FRENOS",
        "bg": "FCE7F3", "fg": "9D174D",
        "items": [
            ("PAS-DEL-GEN","Pastillas de freno delanteras","Todas",1800,6,3,"Par"),
            ("PAS-TRA-GEN","Pastillas de freno traseras","Todas",1600,6,3,"Par"),
            ("ZAP-TRA-GEN","Zapatas de freno traseras","Todas",1400,4,2,"Par"),
            ("DIS-DEL-GEN","Disco de freno delantero","Todas",3500,4,2,"Unidad"),
            ("DIS-TRA-GEN","Disco de freno trasero","Todas",3200,4,2,"Unidad"),
            ("TAM-TRA-GEN","Tambor de freno trasero","Todas",3000,3,2,"Unidad"),
            ("CIL-MAE-GEN","Cilindro maestro de freno","Todas",4500,2,1,"Unidad"),
            ("CIL-RUE-GEN","Cilindro de rueda trasero","Todas",1800,3,2,"Unidad"),
            ("MAN-FRE-DEL","Manguera de freno delantera","Todas",900,4,2,"Unidad"),
            ("MAN-FRE-TRA","Manguera de freno trasera","Todas",850,4,2,"Unidad"),
            ("KIT-CAL-REP","Kit de reparación de mordaza","Todas",1200,4,2,"Kit"),
            ("CAL-DEL-REM","Mordaza de freno delantera remanuf.","Todas",5500,2,1,"Unidad"),
        ]
    },
    {
        "name": "SUSPENSIÓN Y DIRECCIÓN",
        "bg": "EDE9FE", "fg": "5B21B6",
        "items": [
            ("AMO-DEL-GEN","Amortiguador delantero","Todas",4500,4,2,"Unidad"),
            ("AMO-TRA-GEN","Amortiguador trasero","Todas",3800,4,2,"Unidad"),
            ("RES-DEL-GEN","Resorte helicoidal delantero","Todas",3200,2,1,"Unidad"),
            ("RES-TRA-GEN","Resorte helicoidal trasero","Todas",2800,2,1,"Unidad"),
            ("ROT-INF-DEL","Rótula inferior delantera","Todas",2200,4,2,"Unidad"),
            ("ROT-SUP-DEL","Rótula superior delantera","Todas",2200,3,2,"Unidad"),
            ("TER-INT-DIR","Terminal de dirección interior","Todas",1800,4,2,"Unidad"),
            ("TER-EXT-DIR","Terminal de dirección exterior","Todas",1500,4,2,"Unidad"),
            ("BUJ-BAR-EST","Buje barra estabilizadora","Todas",600,6,3,"Unidad"),
            ("LIN-BAR-EST","Link eslabón estabilizadora","Todas",1200,4,2,"Unidad"),
            ("BRA-CON-DEL","Brazo de control delantero A-arm","Todas",5500,2,1,"Unidad"),
            ("BUJ-BRA-CON","Buje de brazo de control","Todas",700,6,3,"Unidad"),
            ("FUE-CRE-DIR","Fuelle cremallera dirección","Todas",1100,4,2,"Unidad"),
            ("CRE-DIR-REM","Cremallera de dirección remanuf.","Todas",12000,1,1,"Unidad"),
            ("KIT-POL-AMO","Kit polvo fuelle amortiguador","Todas",900,4,2,"Kit"),
        ]
    },
    {
        "name": "SISTEMA ELÉCTRICO",
        "bg": "FEF3C7", "fg": "92400E",
        "items": [
            ("BAT-35AH-GEN","Batería 35AH autos pequeños","Todas",6500,3,2,"Unidad"),
            ("BAT-45AH-GEN","Batería 45AH sedán mediano","Todas",7500,3,2,"Unidad"),
            ("BAT-65AH-GEN","Batería 65AH SUV pickup","Todas",9500,2,1,"Unidad"),
            ("ALT-REM-GEN","Alternador remanufacturado","Todas",12000,2,1,"Unidad"),
            ("ARR-REM-GEN","Motor de arranque remanuf.","Todas",9000,2,1,"Unidad"),
            ("SEN-O2-UPS","Sensor de oxígeno upstream","Todas",3500,3,2,"Unidad"),
            ("SEN-O2-DWN","Sensor de oxígeno downstream","Todas",3000,3,2,"Unidad"),
            ("SEN-MAF-GEN","Sensor MAF flujo de masa de aire","Todas",5500,2,1,"Unidad"),
            ("SEN-MAP-GEN","Sensor MAP","Todas",3200,2,1,"Unidad"),
            ("SEN-ECT-GEN","Sensor temperatura refrigerante ECT","Todas",1800,3,2,"Unidad"),
            ("SEN-CKP-GEN","Sensor posición cigüeñal CKP","Todas",2800,3,2,"Unidad"),
            ("SEN-CMP-GEN","Sensor posición árbol levas CMP","Todas",2800,3,2,"Unidad"),
            ("SEN-TPS-GEN","Sensor TPS posición acelerador","Todas",2500,2,1,"Unidad"),
            ("REL-VEN-GEN","Relay módulo control ventilador","Todas",1800,3,2,"Unidad"),
            ("REL-PRI-GEN","Relay principal main relay","Todas",1200,4,2,"Unidad"),
            ("FUS-SUR-GEN","Fusibles surtido caja","Todas",500,5,3,"Caja"),
            ("BOM-H4-HAL","Bombilla H4 halógena","Todas",600,8,4,"Unidad"),
            ("BOM-H7-HAL","Bombilla H7 halógena","Todas",600,8,4,"Unidad"),
            ("BOM-H11-HAL","Bombilla H11 antiniebla","Todas",600,8,4,"Unidad"),
            ("KIT-LED-CON","Kit conversión LED","Todas",2500,4,2,"Kit"),
            ("SWI-STO-GEN","Switch luces de stop","Todas",900,4,2,"Unidad"),
        ]
    },
    {
        "name": "SISTEMA DE ENFRIAMIENTO",
        "bg": "CCFBF1", "fg": "065F46",
        "items": [
            ("TER-GEN-88","Termostato 88°C","Todas",900,5,3,"Unidad"),
            ("BOM-AGU-GEN","Bomba de agua","Todas",3500,3,2,"Unidad"),
            ("MAN-RAD-SUP","Manguera superior radiador","Todas",1200,4,2,"Unidad"),
            ("MAN-RAD-INF","Manguera inferior radiador","Todas",1000,4,2,"Unidad"),
            ("RAD-ALU-GEN","Radiador aluminio universal","Todas",18000,2,1,"Unidad"),
            ("TAP-RAD-09","Tapa de radiador 0.9 bar","Todas",600,5,3,"Unidad"),
            ("TAP-RAD-11","Tapa de radiador 1.1 bar","Todas",600,5,3,"Unidad"),
            ("DEP-EXP-GEN","Depósito de expansión","Todas",2000,3,1,"Unidad"),
            ("VEN-ELE-RAD","Ventilador eléctrico de radiador","Todas",5000,2,1,"Unidad"),
            ("SEN-TEM-RAD","Sensor temperatura radiador fan switch","Todas",1500,3,2,"Unidad"),
        ]
    },
    {
        "name": "MOTOR Y TRANSMISIÓN",
        "bg": "FEE2E2", "fg": "991B1B",
        "items": [
            ("COR-DIS-GEN","Correa de distribución","Todas",2500,4,2,"Unidad"),
            ("TEN-DIS-GEN","Tensor correa distribución","Todas",2000,3,2,"Unidad"),
            ("POL-LOC-DIS","Polea loca de distribución","Todas",1800,3,2,"Unidad"),
            ("KIT-DIS-COM","Kit distribución completo","Todas",7500,3,2,"Kit"),
            ("COR-SER-GEN","Correa serpentina","Todas",1800,4,2,"Unidad"),
            ("TEN-SER-GEN","Tensor correa serpentina","Todas",2200,3,2,"Unidad"),
            ("JUN-CAB-GEN","Junta de cabeza empaque culata","Todas",8500,2,1,"Unidad"),
            ("SEL-CIG-DEL","Sello retén cigüeñal delantero","Todas",700,4,2,"Unidad"),
            ("SEL-CIG-TRA","Sello retén cigüeñal trasero","Todas",700,4,2,"Unidad"),
            ("SEL-ALC-GEN","Sello árbol de levas","Todas",600,4,2,"Unidad"),
            ("VAL-PCV-GEN","Válvula PCV","Todas",800,4,2,"Unidad"),
            ("JUN-CAR-GEN","Junta de cárter de aceite","Todas",600,4,2,"Unidad"),
            ("TAP-DRE-GEN","Tapón de drenaje de aceite","Todas",200,10,5,"Unidad"),
            ("CAD-DIS-GEN","Cadena de distribución","Todas",5500,2,1,"Unidad"),
            ("KIT-CAD-DIS","Kit cadena distribución completo","Todas",14000,2,1,"Kit"),
        ]
    },
    {
        "name": "AIRE ACONDICIONADO",
        "bg": "E0F2FE", "fg": "075985",
        "items": [
            ("COM-AC-REM","Compresor A/C remanufacturado","Todas",22000,1,1,"Unidad"),
            ("FIL-SEC-AC","Filtro secador de A/C","Todas",2200,3,2,"Unidad"),
            ("VAL-EXP-AC","Válvula de expansión","Todas",2800,2,1,"Unidad"),
            ("MAN-AC-ALT","Manguera A/C alta presión","Todas",3500,2,1,"Unidad"),
            ("MAN-AC-BAJ","Manguera A/C baja presión","Todas",3000,2,1,"Unidad"),
            ("GAS-R134-LAT","Gas R134a lata 12oz","Todas",800,10,5,"Unidad"),
            ("GAS-R134-CIL","Gas R134a cilindro 30 lb","Todas",8500,2,1,"Cilindro"),
            ("ACE-COM-POE","Aceite compresor A/C POE 46","Todas",1200,4,2,"Unidad"),
        ]
    },
    {
        "name": "TRANSMISIÓN Y TRACCIÓN",
        "bg": "FFE4E6", "fg": "9F1239",
        "items": [
            ("FUE-HOI-INT","Fuelle homocinética interior","Todas",1500,4,2,"Unidad"),
            ("FUE-HOI-EXT","Fuelle homocinética exterior","Todas",1200,4,2,"Unidad"),
            ("JUN-HOI-EXT","Junta homocinética externa completa","Todas",4500,3,2,"Unidad"),
            ("SEM-EJE-REM","Semieje completo remanuf.","Todas",9500,2,1,"Unidad"),
            ("KIT-SEL-DIF","Kit sellos diferencial","Todas",1200,3,2,"Kit"),
            ("SEN-ABS-RUE","Sensor velocidad ABS por rueda","Todas",2500,4,2,"Unidad"),
        ]
    },
    {
        "name": "SISTEMA DE ESCAPE",
        "bg": "ECFDF5", "fg": "064E3B",
        "items": [
            ("CAT-UNI-GEN","Catalizador universal","Todas",12000,2,1,"Unidad"),
            ("SIL-TRA-GEN","Silenciador mofle trasero","Todas",6000,2,1,"Unidad"),
            ("JUN-ESC-MAN","Junta escape manifold","Todas",900,4,2,"Unidad"),
            ("TUB-INT-ESC","Tubo intermedio de escape","Todas",4500,2,1,"Unidad"),
            ("SEN-TEM-ESC","Sensor temperatura de escape","Todas",2800,2,1,"Unidad"),
        ]
    },
    {
        "name": "CONSUMIBLES DE TALLER",
        "bg": "F3F4F6", "fg": "374151",
        "items": [
            ("ADI-INY-LIM","Aditivo limpia inyectores","Todas",600,8,4,"Unidad"),
            ("ADI-COM-LIM","Aditivo limpia sistema combustible","Todas",500,8,4,"Unidad"),
            ("SEL-RTV-NEG","Sellador RTV silicona negro","Todas",700,6,3,"Tubo"),
            ("SEL-RTV-ROJ","Sellador RTV silicona rojo alta temp","Todas",700,6,3,"Tubo"),
            ("LOC-ROS-GEN","Loctite sellador de roscas","Todas",600,5,3,"Unidad"),
            ("GRA-MUL-GEN","Grasa multipropósito","Todas",400,8,4,"Unidad"),
            ("GRA-CHA-GEN","Grasa de chasis axiales","Todas",500,6,3,"Unidad"),
            ("SPR-CON-ELE","Spray limpia contactos eléctricos","Todas",800,4,2,"Unidad"),
            ("SPR-WD40","Spray afloja todo WD-40","Todas",700,6,3,"Unidad"),
            ("CIN-AIS-CAL","Cinta aislante alta temperatura","Todas",400,8,4,"Rollo"),
            ("ABR-MET-SUR","Abrazaderas metálicas surtido","Todas",350,10,5,"Bolsa"),
        ]
    },
]

# ── SHEET 1: Inventario ───────────────────────────────────────────────────────
ws = wb.active
ws.title = "Inventario"

# Row 1 — Title
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "SÓLIDO AUTO SERVICIO — Inventario de Repuestos"
c.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
c.fill = hex_fill("#1E3A5F")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Row 2 — Brands
ws.merge_cells("A2:I2")
c = ws["A2"]
c.value = "Toyota · Honda · Mazda · Nissan · Chevrolet · Hyundai · Kia · Ford"
c.font = Font(name="Arial", color="ADD8E6", italic=True, size=11)
c.fill = hex_fill("#000000")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Row 3 — Headers
headers = ["Categoría","Código","Nombre del Repuesto","Marcas Compatibles","Precio (RD$)","Stock Actual","Stock Mínimo","Unidad","Estado"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = hex_fill("#000000")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(c)
ws.row_dimensions[3].height = 22

# Column widths
col_widths = [25, 18, 35, 20, 14, 13, 13, 10, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data rows
current_row = 4
summary_data = []

for cat in categories:
    cat_start = current_row
    items_to_restock = 0
    prices = []
    total_stock = 0

    for item in cat["items"]:
        codigo, nombre, marcas, precio, stock_act, stock_min, unidad = item
        row = current_row

        ws.cell(row=row, column=1, value=cat["name"])
        ws.cell(row=row, column=2, value=codigo)
        ws.cell(row=row, column=3, value=nombre)
        ws.cell(row=row, column=4, value=marcas)
        ws.cell(row=row, column=5, value=precio)
        ws.cell(row=row, column=6, value=stock_act)
        ws.cell(row=row, column=7, value=stock_min)
        ws.cell(row=row, column=8, value=unidad)
        ws.cell(row=row, column=9, value=f'=IF(F{row}>=G{row},"✅ OK","⚠️ Reabastecer")')

        # Styles
        row_fill = hex_fill("#" + cat["bg"])
        row_font_color = cat["fg"]

        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.font = Font(name="Arial", size=9, color=row_font_color)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))
            apply_border(cell)

        # Price column: green font, number format
        price_cell = ws.cell(row=row, column=5)
        price_cell.font = Font(name="Arial", size=9, color="006400", bold=True)
        price_cell.number_format = "#,##0"
        price_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Stock cells centered
        for col in [6, 7]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Estado centered
        ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")

        # Categoría col left-aligned
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

        prices.append(precio)
        total_stock += stock_act
        if stock_act < stock_min:
            items_to_restock += 1

        current_row += 1

    summary_data.append({
        "cat": cat["name"],
        "n": len(cat["items"]),
        "avg_price": sum(prices) / len(prices) if prices else 0,
        "total_stock": total_stock,
        "restock": items_to_restock,
    })

last_data_row = current_row - 1

# Freeze panes at A4
ws.freeze_panes = "A4"

# AutoFilter on row 3
ws.auto_filter.ref = f"A3:I{last_data_row}"

# ── SHEET 2: Resumen por Categoría ───────────────────────────────────────────
ws2 = wb.create_sheet("Resumen por Categoría")

# Title
ws2.merge_cells("A1:E1")
c2 = ws2["A1"]
c2.value = "SÓLIDO AUTO SERVICIO — Resumen por Categoría"
c2.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
c2.fill = hex_fill("#1E3A5F")
c2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

# Headers
res_headers = ["Categoría","N° Repuestos","Precio Promedio (RD$)","Stock Total","Items a Reabastecer"]
for col, h in enumerate(res_headers, 1):
    c2 = ws2.cell(row=2, column=col, value=h)
    c2.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c2.fill = hex_fill("#1E3A5F")
    c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(c2)
ws2.row_dimensions[2].height = 22

# Data
for i, s in enumerate(summary_data, 3):
    ws2.cell(row=i, column=1, value=s["cat"]).font = Font(name="Arial", bold=True, size=9)
    ws2.cell(row=i, column=2, value=s["n"]).alignment = Alignment(horizontal="center")
    avg_cell = ws2.cell(row=i, column=3, value=round(s["avg_price"], 0))
    avg_cell.number_format = "#,##0"
    avg_cell.font = Font(name="Arial", size=9, color="006400")
    avg_cell.alignment = Alignment(horizontal="right")
    ws2.cell(row=i, column=4, value=s["total_stock"]).alignment = Alignment(horizontal="center")
    restock_cell = ws2.cell(row=i, column=5, value=s["restock"])
    restock_cell.alignment = Alignment(horizontal="center")
    if s["restock"] > 0:
        restock_cell.font = Font(name="Arial", bold=True, color="CC0000", size=9)

    alt_fill = hex_fill("#EFF6FF") if i % 2 == 1 else hex_fill("#FFFFFF")
    for col in range(1, 6):
        cell = ws2.cell(row=i, column=col)
        cell.fill = alt_fill
        if cell.font is None or cell.font.color is None or cell.font.color.rgb in ("00000000","FF000000"):
            cell.font = Font(name="Arial", size=9)
        apply_border(cell)

# Totals row
tot_row = len(summary_data) + 3
ws2.cell(row=tot_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
ws2.cell(row=tot_row, column=2, value=sum(s["n"] for s in summary_data)).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
ws2.cell(row=tot_row, column=4, value=sum(s["total_stock"] for s in summary_data)).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
ws2.cell(row=tot_row, column=5, value=sum(s["restock"] for s in summary_data)).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
for col in range(1, 6):
    cell = ws2.cell(row=tot_row, column=col)
    cell.fill = hex_fill("#1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_border(cell)
ws2.cell(row=tot_row, column=2).alignment = Alignment(horizontal="center")

# Column widths for summary
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 20

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = r"D:\SOLIDO AUTO SERVICIO\DOCUMENTOS\PROCESO Y PROCEDIMIENTOS\solido-web\Inventario_Repuestos_Solido.xlsx"
wb.save(output_path)

total_items = sum(len(c["items"]) for c in categories)
print(f"Archivo guardado exitosamente en: {output_path}")
print(f"Total de repuestos incluidos: {total_items}")
print(f"Categorias: {len(categories)}")
